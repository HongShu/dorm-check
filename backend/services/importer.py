"""
Excel 导入逻辑（F-01）

字段映射（需求文档 3.1 & 第十一节）：
- 学号 → student_id
- 姓名 → name
- 性别 → gender
- 行政班 → class_name
- 辅导员 → counselor
- 入住楼栋 → building
- 寝室号 → room
- 在校状态 → status

覆盖更新策略：重复学号覆盖现有记录。
"""

from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from typing import Union

from models.student import Student


# Excel 列名 → 字段名（支持中文列名，也支持带括号后缀的列名如"性别(引用数据源表)"）
FIELD_MAP = {
    "学号": "student_id",
    "姓名": "name",
    "性别": "gender",
    "行政班": "class_name",
    "辅导员": "counselor",
    "入住楼栋": "building",
    "寝室号": "room",
    "在校状态": "status",
}


def _match_header(header: str) -> str | None:
    """列名模糊匹配，处理"性别(引用数据源表)"这种带后缀的情况"""
    if header is None:
        return None
    header = header.strip()
    for key in FIELD_MAP:
        if header == key or header.startswith(key):
            return FIELD_MAP[key]
    return None


def import_students_from_excel(db: Session, source: Union[str, BytesIO]) -> dict:
    """
    从 Excel 文件导入学生数据。source 支持文件路径或 BytesIO。
    导入前会清空 students 表。

    Returns:
        {"imported": int, "by_building": {building: count}}
    """
    # 清空现有数据
    db.query(Student).delete()

    wb = load_workbook(source, data_only=True)
    ws = wb.active

    # 读取表头（第一行）建立列索引
    raw_headers = [cell.value for cell in ws[1]]
    col_index: dict[str, int] = {}
    for idx, h in enumerate(raw_headers):
        field = _match_header(h)
        if field:
            col_index[field] = idx

    if not col_index:
        raise ValueError("Excel 表头无法识别，请确认包含学号、姓名等列")

    imported = 0
    by_building: dict[int, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict: dict[str, any] = {}
        for field, idx in col_index.items():
            val = row[idx]
            if field == "building" and val is not None:
                val = int(val)
            elif field == "room" and val is not None:
                val = str(val).strip()
            row_dict[field] = val

        if not row_dict.get("student_id"):
            continue

        student_id = str(row_dict["student_id"]).strip()
        existing = db.get(Student, student_id)
        if existing:
            for key, val in row_dict.items():
                setattr(existing, key, val)
            existing.updated_at = datetime.utcnow()
        else:
            db.add(Student(**row_dict, updated_at=datetime.utcnow()))

        b = row_dict.get("building")
        if b:
            by_building[b] = by_building.get(b, 0) + 1
        imported += 1

    db.commit()
    wb.close()
    return {"imported": imported, "by_building": by_building}
